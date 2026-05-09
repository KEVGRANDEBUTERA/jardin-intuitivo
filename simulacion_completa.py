import pandas as pd
import math
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.utils import get_column_letter

# Función para limpiar la consola según el sistema operativo
def limpiar():
    os.system('cls' if os.name == 'nt' else 'clear')

# Función para guardar cada simulación en una pestaña diferente de Excel con gráficas
def guardar_en_excel(df, nombre_hoja, tipo_grafica='bar'):
    """
    Guarda el dataframe en Excel y agrega una gráfica automáticamente
    tipo_grafica: 'bar' (barras), 'line' (línea), 'pie' (pastel)
    """
    archivo = "Resultados_Simulacion.xlsx"
    
    if not os.path.isfile(archivo):
        df.to_excel(archivo, sheet_name=nombre_hoja, index=False)
    else:
        with pd.ExcelWriter(archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    
    # Agregar gráfica
    agregar_grafica(archivo, nombre_hoja, df, tipo_grafica)
    print(f"\n✓ Datos guardados en la pestaña: {nombre_hoja}")
    print(f"✓ Gráfica agregada: {tipo_grafica.upper()}")

def agregar_grafica(archivo, nombre_hoja, df, tipo_grafica):
    """
    Agrega una gráfica a la pestaña de Excel especificada
    """
    wb = load_workbook(archivo)
    ws = wb[nombre_hoja]
    
    # Determinar las columnas para la gráfica
    columnas_numericas = df.select_dtypes(include=['number']).columns.tolist()
    
    if len(columnas_numericas) < 1:
        return
    
    # Crear referencia de datos
    max_row = len(df) + 1
    col_datos = 2 if len(df.columns) > 1 else 1
    
    if tipo_grafica == 'bar':
        chart = BarChart()
        chart.title = f"Gráfica: {nombre_hoja}"
        chart.x_axis.title = df.columns[0]
        chart.y_axis.title = columnas_numericas[0]
        
        data = Reference(ws, min_col=col_datos, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
    elif tipo_grafica == 'line':
        chart = LineChart()
        chart.title = f"Gráfica: {nombre_hoja}"
        chart.x_axis.title = df.columns[0]
        chart.y_axis.title = columnas_numericas[0]
        
        data = Reference(ws, min_col=col_datos, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
    elif tipo_grafica == 'pie':
        chart = PieChart()
        chart.title = f"Gráfica: {nombre_hoja}"
        
        labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        data = Reference(ws, min_col=col_datos, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
    
    # Posicionar la gráfica en la pestaña
    ws.add_chart(chart, "D2")
    wb.save(archivo)

def mostrar_estadisticas(df, nombre_modelo):
    """
    Muestra estadísticas básicas del dataframe
    """
    print(f"\n{'='*50}")
    print(f"ESTADÍSTICAS: {nombre_modelo}")
    print(f"{'='*50}")
    print(f"Total de registros: {len(df)}")
    
    columnas_numericas = df.select_dtypes(include=['number']).columns.tolist()
    for col in columnas_numericas:
        if col != 'ri':  # No mostrar stats del ri
            print(f"\n{col}:")
            print(f"  Promedio: {df[col].mean():.4f}")
            print(f"  Máximo: {df[col].max():.4f}")
            print(f"  Mínimo: {df[col].min():.4f}")
            print(f"  Desv. Est.: {df[col].std():.4f}")

# --- INICIO DEL PROGRAMA PRINCIPAL ---
def main():
    # Lista de números aleatorios (ri) extraídos de tus ejercicios
    RI_BASE = [0.213, 0.345, 0.021, 0.987, 0.543, 0.4885, 0.3122, 0.3577, 0.2361, 0.6323, 
               0.6394, 0.6084, 0.6829, 0.0651, 0.0246, 0.0376, 0.7462, 0.5346, 0.3833, 0.3408, 
               0.1365, 0.1250, 0.0876, 0.1321, 0.7300, 0.3545, 0.5596, 0.0898, 0.5765, 0.2129]

    print("\n" + "="*60)
    print("     ★ BIENVENIDO AL SISTEMA DE SIMULACIÓN MULTI-MODELO ★")
    print("="*60)
    
    simulaciones_realizadas = []

    while True:
        limpiar()
        print("\n" + "="*60)
        print("          MENÚ PRINCIPAL - SIMULACIONES")
        print("="*60)
        print("\n📊 MODELOS DE DISTRIBUCIÓN:")
        print("  1. Método Transformada Inversa (Demanda de Cepillos)")
        print("  2. Distribución Uniforme (Temperatura Estufa)")
        print("  3. Distribución Poisson (Piezas por Hora)")
        print("  4. Distribución Bernoulli (Probabilidad de Fallas)")
        print("  5. Distribución Exponencial (Tiempo en Banco)")
        print("\n📈 OPCIONES GENERALES:")
        print("  6. Ver todas las simulaciones realizadas")
        print("  7. Generar archivo final y estadísticas")
        print("  8. Limpiar datos y comenzar de nuevo")
        print("  9. Salir")
        print("="*60)
        
        opcion = input("\n➤ Selecciona una opción (1-9): ").strip()

        if opcion == '1':
            print("\n⏳ Ejecutando: Transformada Inversa (Cepillos)...")
            res = [0 if r < 0.1111 else 1 if r < 0.4444 else 2 if r < 0.6666 else 3 for r in RI_BASE]
            df = pd.DataFrame({
                'Día': range(1, 31), 
                'ri': RI_BASE, 
                'Demanda': res
            })
            guardar_en_excel(df, "Transformada_Inversa", 'bar')
            mostrar_estadisticas(df, "Demanda de Cepillos")
            simulaciones_realizadas.append("Transformada Inversa")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '2':
            print("\n⏳ Ejecutando: Distribución Uniforme (Estufa)...")
            temp = [round(95 + 5*r, 2) for r in RI_BASE]
            df = pd.DataFrame({
                'Medición': range(1, 31), 
                'ri': RI_BASE, 
                'Temp_C': temp
            })
            guardar_en_excel(df, "Uniforme", 'line')
            mostrar_estadisticas(df, "Temperatura en Estufa")
            simulaciones_realizadas.append("Distribución Uniforme")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '3':
            print("\n⏳ Ejecutando: Distribución Poisson (Piezas)...")
            def p_val(r):
                if r < 0.1353: return 0
                if r < 0.4060: return 1
                if r < 0.6767: return 2
                if r < 0.8571: return 3
                if r < 0.9473: return 4
                return 5
            res = [p_val(r) for r in RI_BASE[:20]]
            df = pd.DataFrame({
                'Hora': range(1, 21), 
                'ri': RI_BASE[:20], 
                'Piezas': res
            })
            guardar_en_excel(df, "Poisson", 'bar')
            mostrar_estadisticas(df, "Producción de Piezas")
            simulaciones_realizadas.append("Distribución Poisson")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '4':
            print("\n⏳ Ejecutando: Distribución Bernoulli (Fallas)...")
            res = [1 if r <= 0.2 else 0 for r in RI_BASE]
            df = pd.DataFrame({
                'Día': range(1, 31), 
                'ri': RI_BASE, 
                'Falla': res
            })
            guardar_en_excel(df, "Bernoulli", 'pie')
            mostrar_estadisticas(df, "Probabilidad de Fallas")
            simulaciones_realizadas.append("Distribución Bernoulli")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '5':
            print("\n⏳ Ejecutando: Distribución Exponencial (Banco)...")
            tiempos = [round(-3 * math.log(1 - r), 2) for r in RI_BASE[:15]]
            df = pd.DataFrame({
                'Cliente': range(1, 16), 
                'ri': RI_BASE[:15], 
                'Tiempo_min': tiempos
            })
            guardar_en_excel(df, "Exponencial", 'line')
            mostrar_estadisticas(df, "Tiempo de Atención en Banco")
            simulaciones_realizadas.append("Distribución Exponencial")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '6':
            limpiar()
            print("\n" + "="*60)
            print("        SIMULACIONES REALIZADAS EN ESTA SESIÓN")
            print("="*60)
            if simulaciones_realizadas:
                for i, sim in enumerate(simulaciones_realizadas, 1):
                    print(f"  {i}. {sim}")
                print(f"\nTotal: {len(simulaciones_realizadas)} simulación(es)")
            else:
                print("\n⚠️  Aún no has realizado ninguna simulación.")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '7':
            limpiar()
            print("\n" + "="*60)
            print("     FINALIZANDO SIMULACIONES Y GENERANDO REPORTES")
            print("="*60)
            if os.path.isfile("Resultados_Simulacion.xlsx"):
                print("\n✓ Archivo 'Resultados_Simulacion.xlsx' creado correctamente")
                print("\n📊 Simulaciones incluidas:")
                for i, sim in enumerate(simulaciones_realizadas, 1):
                    print(f"  {i}. {sim}")
                print(f"\n✓ Total de hojas: {len(simulaciones_realizadas)}")
                print("\n📁 El archivo se encuentra en la carpeta del proyecto")
                print("   Abre el archivo para ver todas las gráficas generadas")
            else:
                print("\n⚠️  No hay datos guardados aún. Ejecuta primero algunas simulaciones.")
            input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '8':
            confirmacion = input("\n⚠️  ¿Estás seguro de que deseas eliminar todos los datos? (s/n): ").strip().lower()
            if confirmacion == 's':
                if os.path.isfile("Resultados_Simulacion.xlsx"):
                    os.remove("Resultados_Simulacion.xlsx")
                    print("\n✓ Datos eliminados correctamente")
                    simulaciones_realizadas = []
                else:
                    print("\n⚠️  No hay archivos para eliminar")
                input("\n✓ Presiona ENTER para continuar...")
            else:
                print("\n✓ Operación cancelada")
                input("\n✓ Presiona ENTER para continuar...")

        elif opcion == '9':
            limpiar()
            print("\n" + "="*60)
            print("              👋 GRACIAS POR USAR EL SIMULADOR")
            print("="*60)
            if os.path.isfile("Resultados_Simulacion.xlsx"):
                print("\n✓ Revisa el archivo: 'Resultados_Simulacion.xlsx'")
            print("\n")
            break

        else:
            print("\n❌ Error: Opción no válida. Por favor, selecciona una opción entre 1 y 9.")
            input("\n✓ Presiona ENTER para continuar...")

if __name__ == "__main__":
    main()
